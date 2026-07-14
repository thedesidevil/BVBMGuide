from __future__ import annotations
import io
import re

import openpyxl

from src.hotel_options.decoder import decode_col_h
from src.hotel_options.models import (
    HotelRow, PlanPricing, Plan, UnknownCode, ParseResult,
)

_PLAN_RE = re.compile(
    r'^(?:PLAN\s+\w+|.*\bPlan)(?:\s*\(recommended\))?\s*$',
    re.IGNORECASE,
)
_SECTION_DATE_RE = re.compile(r'\(([^)]+)\)')
_RECOMMENDED_RE = re.compile(r'\s*\(recommended\)\s*$', re.IGNORECASE)


def _strip_recommended(s: str) -> tuple[str, bool]:
    """Return (cleaned_string, recommended_flag)."""
    if _RECOMMENDED_RE.search(s):
        return _RECOMMENDED_RE.sub('', s).strip(), True
    return s, False
_FILENAME_RE = re.compile(
    r'(?:copy\s+of\s+)?(?:DO NOT SHARE[\s_]+)?([^_]+)_Accommodation Options_([^_.]+)',
    re.IGNORECASE,
)
_TRAILING_COPY_NUM_RE = re.compile(r'\s*\(\d+\)\s*$')


def extract_filename_meta(filename: str) -> tuple[str, str]:
    m = _FILENAME_RE.search(filename)
    if m:
        client_name = m.group(1).strip()
        destination = _TRAILING_COPY_NUM_RE.sub('', m.group(2)).strip()
        return client_name, destination
    stem = filename.rsplit(".", 1)[0]
    parts = [p.strip() for p in stem.split("_") if p.strip()]
    client_name = ""
    if parts:
        first = parts[0]
        cleaned = re.sub(r'^(?:copy\s+of\s+)?DO NOT SHARE[\s_]*', '', first, flags=re.IGNORECASE).strip()
        if cleaned:
            client_name = cleaned
        elif len(parts) >= 2:
            client_name = parts[1]
    destination = _TRAILING_COPY_NUM_RE.sub('', parts[-1]).strip() if parts else ""
    return client_name, destination


def _numeric(v) -> float | None:
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        clean = v.strip().lstrip('₹$€£¥').replace(',', '').strip()
        try:
            result = float(clean)
            return result if result > 0 else None
        except (ValueError, TypeError):
            return None
    return None


def _make_dummy_row(length: int = 14) -> tuple:
    """Create a synthetic row with all None values for flushing without a real summary."""
    class NoneCell:
        value = None
    return tuple(NoneCell() for _ in range(length))


def _parse_no_plans(ws, codes: dict) -> tuple[list[Plan], list[UnknownCode]]:
    """Group hotels by section headers when no PLAN markers found.
    Returns (plans, unknown_codes). If no section headers, one "All Hotels" plan."""
    all_plans: list[Plan] = []
    unknown_codes: list[UnknownCode] = []
    current_label: str | None = None
    current_hotels: list[HotelRow] = []
    running_online = 0.0
    running_b2b = 0.0

    def _flush():
        nonlocal current_label, current_hotels, running_online, running_b2b
        if not current_hotels:
            current_label = None
            current_hotels = []
            running_online = 0.0
            running_b2b = 0.0
            return
        pricing = PlanPricing(
            total_online_price=running_online,
            total_b2b_price=running_b2b,
            customer_discount=0.0,
            discounted_price=running_online,
            discount_pct=0.0,
        )
        all_plans.append(Plan(
            label=current_label if current_label is not None else "All Hotels",
            hotels=list(current_hotels),
            pricing=pricing,
        ))
        current_label = None
        current_hotels = []
        running_online = 0.0
        running_b2b = 0.0

    for row in ws.iter_rows():
        cell_a = row[0]
        row_dim = ws.row_dimensions.get(cell_a.row)
        if (cell_a.font and cell_a.font.strike) or (row_dim and row_dim.font and row_dim.font.strike):
            continue

        val_a = cell_a.value
        str_a = str(val_a).strip() if val_a is not None else ""
        col_i_val = row[8].value if len(row) > 8 else None

        # Section header: col A non-empty, col I blank, not a PLAN header, not A1
        if val_a and _numeric(col_i_val) is None and not _PLAN_RE.match(str_a) and cell_a.row > 1:
            _flush()
            current_label = str_a
            continue

        # Hotel row: col A non-empty, col I numeric
        if val_a and _numeric(col_i_val) is not None:
            col_h_val = row[7].value if len(row) > 7 else None
            decoded = decode_col_h(str(col_h_val) if col_h_val is not None else None, codes)
            hotel_name, is_recommended = _strip_recommended(str_a)
            for unk in decoded.unknowns:
                unknown_codes.append(UnknownCode(code=unk, hotel_name=hotel_name, plan_label=current_label or ""))
            online = _numeric(col_i_val) or 0.0
            b2b = (_numeric(row[9].value) if len(row) > 9 else None) or 0.0
            col_l = (_numeric(row[11].value) if len(row) > 11 else None) or 0.0
            col_m = _numeric(row[12].value) if len(row) > 12 else None
            col_n = _numeric(row[13].value) if len(row) > 13 else None
            discount = col_l
            discounted = col_m if col_m is not None else (online - discount)
            pct = col_n if col_n is not None else (discount / online * 100 if online else 0.0)
            running_online += online
            running_b2b += b2b
            why = str(row[17].value).strip() if len(row) > 17 and row[17].value else ""
            current_hotels.append(HotelRow(
                name=hotel_name,
                category=str(row[1].value).strip() if row[1].value else "",
                room_type=str(row[2].value).strip() if row[2].value else "",
                cancellation=decoded.cancellation,
                meal_type=decoded.meal_type,
                online_price=online,
                dates="",
                customer_discount=discount,
                discounted_price=discounted,
                discount_pct=pct,
                recommended=is_recommended,
                why_recommend=why,
            ))

    _flush()
    return all_plans, unknown_codes


def parse_excel(xlsx_bytes: bytes, codes: dict[str, str]) -> ParseResult:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    a1 = ws.cell(1, 1).value
    requirements = str(a1).strip() if a1 is not None else ""

    plans: list[Plan] = []
    unknown_codes: list[UnknownCode] = []

    current_label: str | None = None
    current_recommended: bool = False
    current_why: str = ""
    current_hotels: list[HotelRow] = []
    current_section_dates: str = ""
    current_city: str = ""
    running_online = 0.0
    running_b2b = 0.0
    current_plan_inclusions: str = ""

    def _flush(summary_row) -> None:
        nonlocal current_label, current_recommended, current_why, current_hotels, running_online, running_b2b, current_plan_inclusions
        if current_label is None:
            return
        if not current_hotels:
            current_label = None
            current_recommended = False
            current_why = ""
            current_hotels = []
            running_online = 0.0
            running_b2b = 0.0
            current_plan_inclusions = ""
            return
        col_i = _numeric(summary_row[8].value) if len(summary_row) > 8 else None
        col_j = _numeric(summary_row[9].value) if len(summary_row) > 9 else None
        col_l = _numeric(summary_row[11].value) if len(summary_row) > 11 else None
        col_m = _numeric(summary_row[12].value) if len(summary_row) > 12 else None
        col_n = _numeric(summary_row[13].value) if len(summary_row) > 13 else None

        total_online = col_i if col_i is not None else running_online
        total_b2b = col_j if col_j is not None else running_b2b
        discount = col_l or 0.0
        discounted = col_m if col_m is not None else (total_online - discount)
        pct = col_n if col_n is not None else (discount / total_online * 100 if total_online else 0.0)

        plans.append(Plan(
            label=current_label,
            hotels=list(current_hotels),
            pricing=PlanPricing(
                total_online_price=total_online,
                total_b2b_price=total_b2b,
                customer_discount=discount,
                discounted_price=discounted,
                discount_pct=pct,
            ),
            recommended=current_recommended,
            why_recommend=current_why,
            inclusions=current_plan_inclusions,
        ))
        current_label = None
        current_recommended = False
        current_why = ""
        current_hotels = []
        running_online = 0.0
        running_b2b = 0.0
        current_plan_inclusions = ""

    past_first_plan = False
    # Maps (hotel_name, online_price) → date-range / city from pre-amble section headers.
    # Used as fallback when plan sections have no inline section-date header.
    preamble_dates: dict[tuple[str, float], str] = {}
    preamble_cities: dict[tuple[str, float], str] = {}

    for row in ws.iter_rows():
        cell_a = row[0]

        # Skip any row where col A carries strikethrough — covers PLAN headers,
        # hotel rows, and section headers regardless of how the format was applied.
        row_dim = ws.row_dimensions.get(cell_a.row)
        if (cell_a.font and cell_a.font.strike) or (row_dim and row_dim.font and row_dim.font.strike):
            continue

        val_a = cell_a.value
        str_a = str(val_a).strip() if val_a is not None else ""

        if val_a and _PLAN_RE.match(str_a):
            past_first_plan = True
            if current_label is not None:
                _flush(_make_dummy_row())
            raw_label, current_recommended = _strip_recommended(str_a.title())
            current_label = raw_label
            current_why = str(row[17].value).strip() if len(row) > 17 and row[17].value else ""
            current_hotels = []
            current_section_dates = ""
            current_city = ""
            running_online = 0.0
            running_b2b = 0.0
            continue

        if not past_first_plan:
            # Collect (hotel, price) → dates/city from pre-amble so plan rows can use them.
            col_i_val = row[8].value if len(row) > 8 else None
            if val_a and _numeric(col_i_val) is None:
                m = _SECTION_DATE_RE.search(str_a)
                if m:
                    raw = m.group(1).strip()
                    # Ensure space before hyphen: "Jun 28- Jul 4" → "Jun 28 - Jul 4"
                    current_section_dates = re.sub(r'(\S)-', r'\1 -', raw)
                    current_city = re.sub(r'\s*\(.*\)\s*$', '', str_a).strip()
            elif val_a and _numeric(col_i_val) is not None and current_section_dates:
                price = _numeric(col_i_val)
                if price is not None:
                    preamble_dates[(str_a, price)] = current_section_dates
                    preamble_cities[(str_a, price)] = current_city
            continue

        col_i_val = row[8].value if len(row) > 8 else None
        col_l_val = row[11].value if len(row) > 11 else None

        if str_a.lower() == "transfers":
            continue

        # Plan summary: col A blank, col I or col L numeric — or explicit Total row
        is_total_row = str_a.lower() == "total"
        is_blank_summary = not val_a
        if (is_total_row or is_blank_summary) and (
            _numeric(col_i_val) is not None or _numeric(col_l_val) is not None
        ):
            if is_total_row and len(row) > 18 and row[18].value:
                current_plan_inclusions = str(row[18].value).strip()
            _flush(row)
            continue

        # Section header: col A non-empty, col I blank — extract dates and city if present
        if val_a and _numeric(col_i_val) is None:
            m = _SECTION_DATE_RE.search(str_a)
            if m:
                raw = m.group(1).strip()
                current_section_dates = re.sub(r'(\S)-', r'\1 -', raw)
                current_city = re.sub(r'\s*\(.*\)\s*$', '', str_a).strip()
            continue

        # Hotel row: col A non-empty, col I numeric
        if val_a and _numeric(col_i_val) is not None:

            col_h_val = row[7].value if len(row) > 7 else None
            decoded = decode_col_h(str(col_h_val) if col_h_val is not None else None, codes)

            for unk in decoded.unknowns:
                unknown_codes.append(UnknownCode(
                    code=unk,
                    hotel_name=str_a,
                    plan_label=current_label or "",
                ))

            online_raw = _numeric(col_i_val)
            online = online_raw if online_raw is not None else 0.0
            b2b_raw = _numeric(row[9].value) if len(row) > 9 else None
            b2b = b2b_raw if b2b_raw is not None else 0.0
            running_online += online
            running_b2b += b2b

            dates = current_section_dates or preamble_dates.get((str_a, online), "")
            city = current_city or preamble_cities.get((str_a, online), "")
            why = str(row[17].value).strip() if len(row) > 17 and row[17].value else ""
            current_hotels.append(HotelRow(
                name=str_a,
                category=str(row[1].value).strip() if row[1].value else "",
                room_type=str(row[2].value).strip() if row[2].value else "",
                cancellation=decoded.cancellation,
                meal_type=decoded.meal_type,
                online_price=online,
                dates=dates,
                why_recommend=why,
                city=city,
            ))

    # Flush the last open plan — it may have no trailing summary row
    _flush(_make_dummy_row())
    plans = [p for p in plans if p.hotels]

    grouped_by_sections = False
    if not plans:
        plans, extra_codes = _parse_no_plans(ws, codes)
        unknown_codes.extend(extra_codes)
        grouped_by_sections = bool(plans)

    return ParseResult(
        plans=plans,
        unknown_codes=unknown_codes,
        not_found=[],
        requirements=requirements,
        grouped_by_sections=grouped_by_sections,
    )
