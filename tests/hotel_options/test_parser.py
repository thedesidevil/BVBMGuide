import io
import openpyxl
from src.hotel_options.parser import extract_filename_meta, parse_excel

CODES = {"nr": "Non-refundable", "br": "Breakfast included"}


def make_sample_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "PLAN A"
    ws["A2"] = "Hilton London"
    ws["B2"] = "5-Star"
    ws["C2"] = "King Room"
    ws["H2"] = "nr. br"
    ws["I2"] = 50000.0
    ws["J2"] = 45000.0
    # Plan A summary: col A blank, totals in I/J/L/M/N
    ws["I3"] = 50000.0
    ws["J3"] = 45000.0
    ws["L3"] = 3000.0
    ws["M3"] = 47000.0
    ws["N3"] = 6.0
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_extract_filename_meta_standard():
    name, dest = extract_filename_meta("DO NOT SHARE_ Bushan_Accommodation Options_London.xlsx")
    assert name == "Bushan"
    assert dest == "London"


def test_extract_filename_meta_no_plans_filename():
    name, dest = extract_filename_meta("DO NOT SHARE_ Bushan_no plans.xlsx")
    assert name == "Bushan"
    assert dest == "no plans"


def test_extract_filename_meta_fallback():
    _, dest = extract_filename_meta("some_random_file.xlsx")
    assert dest == "file"


def test_parse_returns_one_plan():
    result = parse_excel(make_sample_xlsx(), CODES)
    assert len(result.plans) == 1
    assert result.plans[0].label == "Plan A"


def test_parse_hotel_decoded():
    result = parse_excel(make_sample_xlsx(), CODES)
    hotel = result.plans[0].hotels[0]
    assert hotel.name == "Hilton London"
    assert hotel.cancellation == "Non-refundable"
    assert hotel.meal_type == "Breakfast included"
    assert hotel.online_price == 50000.0


def test_parse_pricing():
    result = parse_excel(make_sample_xlsx(), CODES)
    pricing = result.plans[0].pricing
    assert pricing.total_online_price == 50000.0
    assert pricing.customer_discount == 3000.0
    assert pricing.discounted_price == 47000.0
    assert pricing.discount_pct == 6.0


def test_no_unknown_codes():
    result = parse_excel(make_sample_xlsx(), CODES)
    assert result.unknown_codes == []


def test_unknown_code_flagged():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "PLAN A"
    ws["A2"] = "Test Hotel"
    ws["B2"] = "3-Star"
    ws["C2"] = "Double"
    ws["H2"] = "xy"
    ws["I2"] = 10000.0
    ws["J2"] = 9000.0
    ws["I3"] = 10000.0
    ws["J3"] = 9000.0
    ws["L3"] = 0.0
    ws["M3"] = 10000.0
    ws["N3"] = 0.0
    buf = io.BytesIO()
    wb.save(buf)
    result = parse_excel(buf.getvalue(), CODES)
    assert len(result.unknown_codes) == 1
    assert result.unknown_codes[0].code == "xy"


def test_last_plan_no_trailing_summary_is_included():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "PLAN A"
    ws["A2"] = "Hotel Alpha"; ws["B2"] = "3-Star"; ws["C2"] = "Twin"; ws["I2"] = 10000.0; ws["J2"] = 9000.0
    ws["I3"] = 10000.0; ws["J3"] = 9000.0; ws["L3"] = 500.0; ws["M3"] = 9500.0; ws["N3"] = 5.0
    ws["A4"] = "PLAN B"
    ws["A5"] = "Hotel Beta"; ws["B5"] = "4-Star"; ws["C5"] = "Double"; ws["I5"] = 20000.0; ws["J5"] = 18000.0
    # No summary row for Plan B — sheet ends here
    buf = io.BytesIO(); wb.save(buf); xlsx = buf.getvalue()
    result = parse_excel(xlsx, {})
    assert len(result.plans) == 2
    assert result.plans[1].label == "Plan B"
    assert result.plans[1].hotels[0].name == "Hotel Beta"


def test_plans_present_grouped_by_sections_false():
    """Existing PLAN-marker workbooks should have grouped_by_sections=False."""
    result = parse_excel(make_sample_xlsx(), CODES)
    assert result.grouped_by_sections is False


# Helper to build a minimal in-memory workbook for testing
def _make_wb_no_plans_with_sections():
    """Two section headers, two hotels each, no PLAN markers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    # Row 1: requirements in A1
    ws.cell(1, 1).value = "Breakfast Included"
    # Row 2: section header  (col A text, col I blank)
    ws.cell(2, 1).value = "London (Jul 1 - Jul 5)"
    # Row 3 & 4: hotel rows under London
    ws.cell(3, 1).value = "Hotel Alpha"
    ws.cell(3, 2).value = "4-Star"
    ws.cell(3, 3).value = "Deluxe"
    ws.cell(3, 8).value = "nr"
    ws.cell(3, 9).value = 50000.0
    ws.cell(3, 10).value = 45000.0
    ws.cell(4, 1).value = "Hotel Beta"
    ws.cell(4, 2).value = "3-Star"
    ws.cell(4, 3).value = "Standard"
    ws.cell(4, 8).value = "br"
    ws.cell(4, 9).value = 40000.0
    ws.cell(4, 10).value = 36000.0
    # Row 5: section header
    ws.cell(5, 1).value = "Paris (Jul 6 - Jul 10)"
    # Row 6: hotel row under Paris
    ws.cell(6, 1).value = "Hotel Gamma"
    ws.cell(6, 2).value = "5-Star"
    ws.cell(6, 3).value = "Suite"
    ws.cell(6, 8).value = "nr"
    ws.cell(6, 9).value = 80000.0
    ws.cell(6, 10).value = 72000.0
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_wb_no_plans_flat():
    """Hotels with no section headers and no PLAN markers — one group."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1).value = "Any requirement"
    for i, name in enumerate(["Hotel X", "Hotel Y", "Hotel Z"], start=2):
        ws.cell(i, 1).value = name
        ws.cell(i, 2).value = "4-Star"
        ws.cell(i, 3).value = "Double"
        ws.cell(i, 8).value = "nr"
        ws.cell(i, 9).value = 30000.0
        ws.cell(i, 10).value = 27000.0
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_no_plans_with_sections_groups_by_section():
    result = parse_excel(_make_wb_no_plans_with_sections(), codes={})
    assert result.grouped_by_sections is True
    assert len(result.plans) == 2
    london = result.plans[0]
    paris = result.plans[1]
    assert london.label == "London (Jul 1 - Jul 5)"  # raw label before normalization
    assert paris.label == "Paris (Jul 6 - Jul 10)"
    assert len(london.hotels) == 2
    assert len(paris.hotels) == 1
    assert london.pricing.total_online_price == 90000.0
    assert paris.pricing.total_online_price == 80000.0
    assert london.pricing.customer_discount == 0.0
    assert london.pricing.discounted_price == 90000.0
    assert london.pricing.discount_pct == 0.0


def test_no_plans_flat_all_hotels_in_one_plan():
    result = parse_excel(_make_wb_no_plans_flat(), codes={})
    assert result.grouped_by_sections is True
    assert len(result.plans) == 1
    assert result.plans[0].label == "All Hotels"
    assert len(result.plans[0].hotels) == 3
    assert result.plans[0].pricing.total_online_price == 90000.0
    assert result.plans[0].pricing.customer_discount == 0.0
    assert result.plans[0].pricing.discounted_price == 90000.0
    assert result.plans[0].pricing.discount_pct == 0.0


def test_no_plans_per_hotel_pricing():
    """Cols L/M/N are captured as per-hotel discount data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1).value = "Breakfast included"
    ws.cell(2, 1).value = "London (Jul 1 - Jul 4)"
    ws.cell(3, 1).value = "Hotel Alpha"
    ws.cell(3, 2).value = "4-Star"
    ws.cell(3, 3).value = "Double"
    ws.cell(3, 8).value = "nr"
    ws.cell(3, 9).value = 100000.0   # col I online
    ws.cell(3, 10).value = 90000.0   # col J b2b
    ws.cell(3, 12).value = 5000.0    # col L customer_discount
    ws.cell(3, 13).value = 95000.0   # col M discounted_price
    ws.cell(3, 14).value = 5.0       # col N discount_pct
    buf = io.BytesIO()
    wb.save(buf)
    result = parse_excel(buf.getvalue(), codes={})
    assert len(result.plans) == 1
    hotel = result.plans[0].hotels[0]
    assert hotel.online_price == 100000.0
    assert hotel.customer_discount == 5000.0
    assert hotel.discounted_price == 95000.0
    assert hotel.discount_pct == 5.0


def test_no_plans_string_prices_in_col_i():
    """Col I contains string-formatted currency values (e.g. '1,50,000' or '₹50,000')."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1).value = "2 adults\nrefundable with breakfast"
    ws.cell(2, 1).value = "London"
    ws.cell(3, 1).value = "Hotel Alpha"
    ws.cell(3, 2).value = "4-Star"
    ws.cell(3, 3).value = "Deluxe"
    ws.cell(3, 8).value = "nr"
    ws.cell(3, 9).value = "1,50,000"   # Indian thousands-separated string
    ws.cell(4, 1).value = "Hotel Beta"
    ws.cell(4, 2).value = "3-Star"
    ws.cell(4, 3).value = "Standard"
    ws.cell(4, 8).value = "br"
    ws.cell(4, 9).value = "₹2,00,000"  # With rupee symbol
    buf = io.BytesIO()
    wb.save(buf)
    result = parse_excel(buf.getvalue(), codes={})
    assert result.grouped_by_sections is True
    assert len(result.plans) == 1
    assert result.plans[0].label == "London"
    assert len(result.plans[0].hotels) == 2
    assert result.plans[0].hotels[0].online_price == 150000.0
    assert result.plans[0].hotels[1].online_price == 200000.0
    assert result.plans[0].pricing.total_online_price == 350000.0


def test_recommended_flag_on_hotel_no_plans():
    """'(Recommended)' suffix in col A sets recommended=True and is stripped from name."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1).value = "Breakfast included"
    ws.cell(2, 1).value = "London"
    ws.cell(3, 1).value = "Hotel Alpha (Recommended)"
    ws.cell(3, 2).value = "4-Star"; ws.cell(3, 3).value = "Double"
    ws.cell(3, 8).value = "nr"; ws.cell(3, 9).value = 50000.0
    ws.cell(4, 1).value = "Hotel Beta"
    ws.cell(4, 2).value = "3-Star"; ws.cell(4, 3).value = "Standard"
    ws.cell(4, 8).value = "br"; ws.cell(4, 9).value = 40000.0
    buf = io.BytesIO(); wb.save(buf)
    result = parse_excel(buf.getvalue(), codes={})
    hotels = result.plans[0].hotels
    assert hotels[0].name == "Hotel Alpha"
    assert hotels[0].recommended is True
    assert hotels[1].recommended is False


def test_recommended_flag_on_plan():
    """'(Recommended)' in PLAN header sets plan.recommended=True and is stripped from label."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "PLAN A (Recommended)"
    ws["A2"] = "Hilton London"; ws["B2"] = "5-Star"; ws["C2"] = "King"
    ws["H2"] = "nr"; ws["I2"] = 50000.0; ws["J2"] = 45000.0
    ws["I3"] = 50000.0; ws["J3"] = 45000.0; ws["L3"] = 3000.0; ws["M3"] = 47000.0; ws["N3"] = 6.0
    ws["A4"] = "PLAN B"
    ws["A5"] = "Marriott"; ws["B5"] = "4-Star"; ws["C5"] = "Double"
    ws["H5"] = "nr"; ws["I5"] = 60000.0; ws["J5"] = 55000.0
    ws["I6"] = 60000.0; ws["J6"] = 55000.0; ws["L6"] = 0.0; ws["M6"] = 60000.0; ws["N6"] = 0.0
    buf = io.BytesIO(); wb.save(buf)
    result = parse_excel(buf.getvalue(), codes={})
    assert result.plans[0].label == "Plan A"
    assert result.plans[0].recommended is True
    assert result.plans[1].label == "Plan B"
    assert result.plans[1].recommended is False


def test_no_plans_empty_section_between_real_sections():
    """A section header with no hotels before the next header is silently dropped."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1).value = "Any requirement"
    # Section 1: "Empty" — no hotels
    ws.cell(2, 1).value = "Empty Section"
    # Section 2: "Active" — two hotels
    ws.cell(3, 1).value = "Active Section"
    ws.cell(4, 1).value = "Hotel One"
    ws.cell(4, 2).value = "4-Star"
    ws.cell(4, 3).value = "Double"
    ws.cell(4, 8).value = "nr"
    ws.cell(4, 9).value = 50000.0
    ws.cell(4, 10).value = 45000.0
    ws.cell(5, 1).value = "Hotel Two"
    ws.cell(5, 2).value = "3-Star"
    ws.cell(5, 3).value = "Standard"
    ws.cell(5, 8).value = "br"
    ws.cell(5, 9).value = 40000.0
    ws.cell(5, 10).value = 36000.0
    buf = io.BytesIO()
    wb.save(buf)
    result = parse_excel(buf.getvalue(), codes={})
    assert result.grouped_by_sections is True
    assert len(result.plans) == 1
    assert result.plans[0].label == "Active Section"
    assert len(result.plans[0].hotels) == 2
    assert result.plans[0].pricing.total_online_price == 90000.0


def _make_star_plan_xlsx() -> bytes:
    """Two plans named '4-Star Plan' and '5-Star Plan' with a Transfers row and Total summary."""
    wb = openpyxl.Workbook()
    ws = wb.active
    # Plan 1
    ws["A1"] = "4-Star Plan"
    ws["A2"] = "Adaaran Resort"
    ws["B2"] = "4-Star"
    ws["C2"] = "Beach Villa"
    ws["H2"] = "Free cancellation before 23 Nov, All Inclusive"
    ws["I2"] = 1000000.0
    ws["A3"] = "Transfers"
    ws["I3"] = 50000.0       # should be ignored
    ws["A4"] = "Total"
    ws["I4"] = 1000000.0
    ws["J4"] = 900000.0
    ws["L4"] = 50000.0
    ws["M4"] = 950000.0
    ws["N4"] = 5.0
    ws["S4"] = "With airport transfer - Shared Speedboat"
    # Plan 2
    ws["A6"] = "5-Star Plan"
    ws["A7"] = "Centara Grand"
    ws["B7"] = "5-Star"
    ws["C7"] = "Overwater Villa"
    ws["H7"] = "nr. br"
    ws["I7"] = 2000000.0
    ws["A8"] = "Total"
    ws["I8"] = 2000000.0
    ws["J8"] = 1800000.0
    ws["L8"] = 100000.0
    ws["M8"] = 1900000.0
    ws["N8"] = 5.0
    ws["S8"] = "With airport transfer - Seaplane"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_star_plan_headers_parsed():
    result = parse_excel(_make_star_plan_xlsx(), {"nr": "Non-refundable", "br": "Breakfast included"})
    assert len(result.plans) == 2
    assert result.plans[0].label == "4-Star Plan"
    assert result.plans[1].label == "5-Star Plan"


def test_transfers_row_skipped():
    result = parse_excel(_make_star_plan_xlsx(), {})
    hotels = result.plans[0].hotels
    assert len(hotels) == 1
    assert hotels[0].name == "Adaaran Resort"


def test_total_row_triggers_plan_flush():
    result = parse_excel(_make_star_plan_xlsx(), {})
    assert len(result.plans) == 2
    assert result.plans[0].pricing.total_online_price == 1000000.0
    assert result.plans[0].pricing.customer_discount == 50000.0
    assert result.plans[0].pricing.discounted_price == 950000.0
    assert result.plans[0].pricing.discount_pct == 5.0


def test_plan_inclusions_from_total_row():
    result = parse_excel(_make_star_plan_xlsx(), {})
    assert result.plans[0].inclusions == "With airport transfer - Shared Speedboat"
    assert result.plans[1].inclusions == "With airport transfer - Seaplane"


def _make_multi_segment_xlsx() -> bytes:
    """One plan, same hotel name with two date-segments."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "4-Star Plan"
    ws["A2"] = "Adaaran Resort (Dec 22-25)"
    ws["B2"] = "4-Star"
    ws["C2"] = "Beach Villa"
    ws["H2"] = "Free cancellation, All Inclusive"
    ws["I2"] = 1178668.0
    ws["S2"] = "Airport transfer"
    ws["T2"] = "Visa fees"
    ws["A3"] = "Adaaran Resort (Dec 25-26)"
    ws["B3"] = "4-Star"
    ws["C3"] = "Ocean Villa"
    ws["H3"] = "Free cancellation, All Inclusive"
    ws["I3"] = 451983.0
    ws["A4"] = "Total"
    ws["I4"] = 1630651.0
    ws["J4"] = 1400000.0
    ws["L4"] = 80000.0
    ws["M4"] = 1550651.0
    ws["N4"] = 4.9
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_hotel_name_date_stripped():
    result = parse_excel(_make_multi_segment_xlsx(), {})
    hotels = result.plans[0].hotels
    assert hotels[0].name == "Adaaran Resort"
    assert hotels[1].name == "Adaaran Resort"


def test_hotel_inline_dates_extracted():
    result = parse_excel(_make_multi_segment_xlsx(), {})
    hotels = result.plans[0].hotels
    assert hotels[0].dates == "Dec 22-25"
    assert hotels[1].dates == "Dec 25-26"


def test_hotel_inclusions_exclusions_from_cols():
    result = parse_excel(_make_multi_segment_xlsx(), {})
    h = result.plans[0].hotels[0]
    assert h.inclusions == "Airport transfer"
    assert h.exclusions == "Visa fees"


def test_extract_filename_meta_dash_separated():
    name, dest = extract_filename_meta("DO NOT SHARE- Vinay- Maldives.xlsx")
    assert name == "Vinay"
    assert dest == "Maldives"


def test_extract_filename_meta_dash_without_do_not_share():
    name, dest = extract_filename_meta("Alice- Greece.xlsx")
    assert name == "Alice"
    assert dest == "Greece"


def test_no_plans_section_label_ending_in_plan_is_section_header():
    """Section label like 'Romantic Plan' must not be swallowed by _PLAN_RE in _parse_no_plans."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1).value = "Breakfast included"
    # Section header: col A text ending in "Plan", col I blank
    ws.cell(2, 1).value = "Romantic Plan"
    ws.cell(3, 1).value = "Hotel Alpha"
    ws.cell(3, 2).value = "4-Star"
    ws.cell(3, 3).value = "Deluxe"
    ws.cell(3, 8).value = "nr"
    ws.cell(3, 9).value = 50000.0
    ws.cell(4, 1).value = "Hotel Beta"
    ws.cell(4, 2).value = "3-Star"
    ws.cell(4, 3).value = "Standard"
    ws.cell(4, 8).value = "br"
    ws.cell(4, 9).value = 40000.0
    buf = io.BytesIO()
    wb.save(buf)
    result = parse_excel(buf.getvalue(), codes={})
    assert result.grouped_by_sections is True
    assert len(result.plans) == 1
    assert result.plans[0].label == "Romantic Plan"
    assert len(result.plans[0].hotels) == 2


def test_preamble_transfers_row_ignored():
    """Transfers row in preamble (before first plan, with numeric price) is skipped."""
    wb = openpyxl.Workbook()
    ws = wb.active
    # Preamble section header with dates
    ws.cell(1, 1).value = "Maldives (Dec 22 - Dec 28)"
    # Transfers row in preamble — col I has a price; must NOT enter preamble_dates
    ws.cell(2, 1).value = "Transfers"
    ws.cell(2, 9).value = 50000.0
    # Real hotel row in preamble
    ws.cell(3, 1).value = "Adaaran Resort"
    ws.cell(3, 9).value = 1000000.0
    # Plan marker
    ws.cell(4, 1).value = "4-Star Plan"
    ws.cell(5, 1).value = "Adaaran Resort"
    ws.cell(5, 2).value = "4-Star"
    ws.cell(5, 3).value = "Beach Villa"
    ws.cell(5, 9).value = 1000000.0
    ws.cell(5, 10).value = 900000.0
    ws.cell(6, 1).value = "Total"
    ws.cell(6, 9).value = 1000000.0
    ws.cell(6, 10).value = 900000.0
    ws.cell(6, 12).value = 0.0
    ws.cell(6, 13).value = 1000000.0
    ws.cell(6, 14).value = 0.0
    buf = io.BytesIO()
    wb.save(buf)
    result = parse_excel(buf.getvalue(), codes={})
    # Transfers must not appear as a hotel
    all_names = [h.name for p in result.plans for h in p.hotels]
    assert "Transfers" not in all_names
    # Real hotel still gets dates from preamble
    assert result.plans[0].hotels[0].dates == "Dec 22 - Dec 28"
