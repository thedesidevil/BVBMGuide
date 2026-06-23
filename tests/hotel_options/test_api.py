import io
import json
from unittest.mock import patch, MagicMock
import openpyxl
import pytest
from fastapi.testclient import TestClient
from src.library.ui import create_app
from src.library.ui.storage import LocalStorageBackend

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def make_xlsx_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "PLAN A"
    ws["A2"] = "Hilton London"
    ws["B2"] = "5-Star"
    ws["C2"] = "King Room"
    ws["H2"] = "nr. br"
    ws["I2"] = 50000.0
    ws["J2"] = 45000.0
    ws["I3"] = 50000.0
    ws["J3"] = 45000.0
    ws["L3"] = 3000.0
    ws["M3"] = 47000.0
    ws["N3"] = 6.0
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def client(tmp_path):
    storage = LocalStorageBackend(tmp_path)
    app = create_app(storage_backend=storage)
    return TestClient(app)


def test_parse_returns_plans(client):
    with patch("src.hotel_options.enricher.check_hotels_exist",
               return_value={"Hilton London": "ChIJ123"}):
        resp = client.post(
            "/api/hotel-options/parse",
            files={"file": ("Bushan_Accommodation Options_London.xlsx",
                            make_xlsx_bytes(), XLSX_MIME)},
        )
    assert resp.status_code == 200
    data = resp.json()
    assert data["client_name"] == "Bushan"
    assert data["destination"] == "London"
    assert len(data["plans"]) == 1
    assert data["plans"][0]["label"] == "Plan A"
    assert data["not_found"] == []


def test_parse_flags_not_found(client):
    with patch("src.hotel_options.enricher.check_hotels_exist",
               return_value={"Hilton London": None}):
        resp = client.post(
            "/api/hotel-options/parse",
            files={"file": ("Bushan_Accommodation Options_London.xlsx",
                            make_xlsx_bytes(), XLSX_MIME)},
        )
    data = resp.json()
    assert len(data["not_found"]) == 1
    assert data["not_found"][0]["sheet_name"] == "Hilton London"


def test_save_code(client):
    resp = client.post("/api/hotel-options/codes", json={"code": "hb", "meaning": "Half board"})
    assert resp.status_code == 200
    assert resp.json() == {"ok": True}


def test_generate_returns_docx(client):
    mock_enriched = MagicMock()
    mock_enriched.official_name = "Hilton London"
    mock_enriched.address = "London"
    mock_enriched.phone = ""
    mock_enriched.rating = 4.0
    mock_enriched.rating_count = 100
    mock_enriched.maps_url = "https://maps.google.com"
    mock_enriched.photo_bytes = None
    mock_enriched.description = "Nice hotel."
    mock_enriched.cancellation = "Non-refundable"
    mock_enriched.meal_type = "Breakfast included"
    mock_enriched.category = "5-Star"

    with (
        patch("src.hotel_options.enricher.check_hotels_exist",
              return_value={"Hilton London": "ChIJ123"}),
        patch("src.hotel_options.enricher.enrich_hotel", return_value=mock_enriched),
    ):
        resp = client.post(
            "/api/hotel-options/generate",
            data={"resolved_codes": "{}", "overrides": "{}"},
            files={"file": ("Bushan_Accommodation Options_London.xlsx",
                            make_xlsx_bytes(), XLSX_MIME)},
        )
    assert resp.status_code == 200
    assert "openxmlformats" in resp.headers.get("content-type", "")
