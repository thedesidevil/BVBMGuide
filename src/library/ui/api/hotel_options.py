from __future__ import annotations
import json
import os

from fastapi import APIRouter, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, Response

from src.hotel_options.codes import CodeStore
from src.library.ui.services.hotel_options_service import parse_file, generate_doc

router = APIRouter()

_XLSX_MAGIC = b"PK"


def _api_key() -> str:
    key = os.getenv("GOOGLE_MAPS_API_KEY", "")
    if not key:
        raise HTTPException(status_code=500, detail="GOOGLE_MAPS_API_KEY not configured")
    return key


@router.post("/hotel-options/parse")
async def parse_hotel_options(request: Request, file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")
    content = await file.read()
    if len(content) < 2 or content[:2] != _XLSX_MAGIC:
        raise HTTPException(status_code=400, detail="File does not appear to be a valid Excel file")
    api_key = _api_key()
    try:
        return parse_file(content, file.filename, request.app.state.storage_backend, api_key)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Parse failed: {e}")


@router.post("/hotel-options/generate")
async def generate_hotel_options(
    request: Request,
    file: UploadFile = File(...),
    resolved_codes: str = Form(default="{}"),
    overrides: str = Form(default="{}"),
):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")
    content = await file.read()
    if len(content) < 2 or content[:2] != _XLSX_MAGIC:
        raise HTTPException(status_code=400, detail="File does not appear to be a valid Excel file")
    try:
        codes_dict = json.loads(resolved_codes)
        overrides_dict = json.loads(overrides)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"Invalid JSON in form fields: {e}")
    api_key = _api_key()
    try:
        docx_bytes, ai_cost_usd, maps_calls = generate_doc(
            content, file.filename, codes_dict, overrides_dict,
            request.app.state.storage_backend, api_key,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Generation failed: {e}")
    return Response(
        content=docx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={
            "Content-Disposition": "attachment; filename=hotel_options.docx",
            "X-AI-Cost-USD": str(round(ai_cost_usd, 4)) if ai_cost_usd is not None else "0",
            "X-Maps-API-Calls": str(maps_calls),
            "Access-Control-Expose-Headers": "X-AI-Cost-USD, X-Maps-API-Calls",
        },
    )


@router.get("/hotel-options/debug-structure", response_class=HTMLResponse)
async def debug_hotel_structure_form():
    return """<html><body style="font-family:monospace;padding:20px">
<h2>Debug: Excel Structure Inspector</h2>
<form method="post" enctype="multipart/form-data">
  <input type="file" name="file" accept=".xlsx" required>
  <button type="submit">Inspect</button>
</form></body></html>"""


@router.post("/hotel-options/debug-structure")
async def debug_hotel_structure(file: UploadFile = File(...)):
    """Return raw cell values from first 25 rows to diagnose parser issues."""
    import io
    import openpyxl
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        table_rows = ""
        for row in ws.iter_rows(max_row=30):
            cells = [(chr(ord('A') + i) if i < 26 else f"col{i}", cell.value, type(cell.value).__name__)
                     for i, cell in enumerate(row) if cell.value is not None]
            if cells:
                cols = " | ".join(f"<b>{c}</b>=<code>{repr(v)}</code> <small>({t})</small>" for c, v, t in cells)
                table_rows += f"<tr><td style='padding:2px 8px;border:1px solid #ccc'>row {row[0].row}</td><td style='padding:2px 8px;border:1px solid #ccc'>{cols}</td></tr>"
        return HTMLResponse(f"""<html><body style="font-family:monospace;padding:20px">
<h2>Structure: {file.filename}</h2>
<p>Dimensions: {ws.max_row} rows × {ws.max_column} cols</p>
<table style="border-collapse:collapse;font-size:13px">{table_rows}</table>
<br><a href="/hotel-options/debug-structure">← Upload another</a>
</body></html>""")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Debug failed: {e}")


@router.post("/hotel-options/codes")
async def save_hotel_code(request: Request, payload: dict):
    code = str(payload.get("code", "")).strip()
    meaning = str(payload.get("meaning", "")).strip()
    if not code or not meaning:
        raise HTTPException(status_code=400, detail="code and meaning are required")
    store = CodeStore(request.app.state.storage_backend)
    codes = store.load()
    codes[code.lower()] = meaning
    store.save(codes)
    return {"ok": True}
